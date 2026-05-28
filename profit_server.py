"""
本地服务模式 - 服务端解析所有文件,无需浏览器 CDN
"""
import requests, json, os, webbrowser, csv, io, re, traceback
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import xlrd
import openpyxl

# === 密钥从同级 config.json 读取，不硬编码 ===
_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")
if not os.path.exists(_CONFIG_PATH):
    print(f"\u2757 缺少配置文件: {_CONFIG_PATH}")
    print("   请复制 config.template.json 为 config.json，填入密钥后重试")
    exit(1)
with open(_CONFIG_PATH, encoding="utf-8") as _f:
    _cfg = json.load(_f)
APP_ID = _cfg["app_id"]
APP_SECRET = _cfg["app_secret"]
COST_SHEET_TOKEN = _cfg["cost_sheet_token"]
PORT = 18632

# ========== 获取飞书成本数据 ==========
def fetch_cost():
    r = requests.post("https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": APP_ID, "app_secret": APP_SECRET})
    token = r.json().get("tenant_access_token")
    headers = {"Authorization": f"Bearer {token}"}

    r = requests.get(f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{COST_SHEET_TOKEN}/values/28e057!A1:CV3550", headers=headers)
    vals = r.json().get("data", {}).get("valueRange", {}).get("values", [])
    putian = []
    for row in vals:
        if len(row) < 96: continue
        code = str(row[0]).strip() if row[0] else ''
        cost_str = str(row[95]).strip() if row[95] else ''
        if code and cost_str:
            try:
                cost = float(cost_str)
                if cost > 0: putian.append([code, cost])
            except: pass

    r = requests.get(f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{COST_SHEET_TOKEN}/values/dOqmQ8!A1:I650", headers=headers)
    vals = r.json().get("data", {}).get("valueRange", {}).get("values", [])
    yiwu = []
    for row in vals[1:]:
        code = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        avg = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        if code and avg:
            try:
                cost = float(avg)
                if cost > 0: yiwu.append([code, cost])
            except: pass
    return {"p": putian, "y": yiwu}

# ========== 通用文件解析器（自动识别 CSV / XLS / XLSX）==========
def parse_any_file(file_data):
    """检测文件魔数，自动选择合适的解析器返回 [{col:val}]"""
    header = file_data[:8]

    # .xls（CFB/OLE2格式）
    if header[:4] == b"\xd0\xcf\x11\xe0":
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        tmp.write(file_data)
        tmp.close()
        try:
            wb = xlrd.open_workbook(tmp.name)
            ws = wb.sheet_by_index(0)
            h = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            rows = []
            for r in range(1, ws.nrows):
                row = {}
                for c in range(ws.ncols):
                    v = ws.cell_value(r, c)
                    row[h[c]] = v if isinstance(v, str) else str(v) if v is not None else ""
                rows.append(row)
            return rows
        finally:
            os.unlink(tmp.name)

    # .xlsx（OOXML / PK ZIP格式）
    if header[:2] == b"PK":
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(file_data)
        tmp.close()
        try:
            wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            ws = wb.active
            h = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                row = {}
                for i, v in enumerate(r):
                    if i < len(h):
                        row[h[i]] = str(v) if v is not None else ""
                rows.append(row)
            wb.close()
            return rows
        finally:
            os.unlink(tmp.name)

    # 默认：作为 CSV 解析
    text = file_data.decode("utf-8-sig")
    text = text.lstrip("\ufeff")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        cleaned = {}
        for k, v in row.items():
            if k is None:
                continue
            key = k.strip().lstrip("\ufeff")
            cleaned[key] = (v or "").strip()
        rows.append(cleaned)
    return rows

# ========== HTML 页面(无 CDN,无外部依赖) ==========
HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>链接盈亏分析工具</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#f5f5f5;color:#333;padding:20px}
.container{max-width:1400px;margin:0 auto}
h1{font-size:22px;margin-bottom:8px}
.ua{background:#fff;border-radius:10px;padding:20px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,0.08)}
.ci{background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:10px 14px;margin-bottom:16px;font-size:13px;color:#0369a1;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.ci .b{display:inline-block;background:#0369a1;color:#fff;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.ur{display:flex;gap:20px;flex-wrap:wrap}
.ui{flex:1;min-width:250px}
.ui label{display:block;font-size:14px;font-weight:600;margin-bottom:5px;color:#333}
.ui .h{font-size:11px;color:#94a3b8;margin-bottom:4px}
.ui input[type=file]{width:100%;padding:7px;border:1px dashed #ccc;border-radius:6px;background:#fafafa;font-size:13px;cursor:pointer}
.ui .s{font-size:12px;margin-top:3px;color:#999}
.br{margin-top:16px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.btn{padding:10px 28px;border:none;border-radius:6px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s;font-family:inherit}
.bp{background:#2563eb;color:#fff}.bp:hover{background:#1d4ed8}
.bp:disabled{background:#93b4f5;cursor:not-allowed}
.bo{background:#fff;color:#555;border:1px solid #ddd}.bo:hover{background:#f0f0f0}
.sum{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin-bottom:16px}
.sc{background:#fff;border-radius:8px;padding:12px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,0.06)}
.sc .n{font-size:20px;font-weight:700}.sc .l{font-size:11px;color:#888;margin-top:3px}
.sc.g .n{color:#22c55e}.sc.r .n{color:#ef4444}.sc.b .n{color:#2563eb}
.cs{background:#fff;border-radius:10px;padding:14px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,0.08)}
.cg{display:grid;grid-template-columns:repeat(auto-fit,minmax(110px,1fr));gap:8px}
.ci2{padding:8px 12px;border-radius:6px;background:#f8fafc;text-align:center}
.ci2 .v{font-size:16px;font-weight:700}.ci2 .p{font-size:10px;color:#888}.ci2 .l{font-size:10px;color:#555;margin-top:2px}
.wb{background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:10px 14px;margin-bottom:12px;font-size:13px;color:#92400e}
.wb.g{background:#f0fdf4;border-color:#86efac;color:#166534}
.tw{background:#fff;border-radius:10px;padding:12px;overflow-x:auto;box-shadow:0 1px 4px rgba(0,0,0,0.08)}
table{width:100%;border-collapse:collapse;font-size:12px;white-space:nowrap}
th{background:#f8fafc;padding:6px 7px;text-align:right;font-weight:600;border-bottom:2px solid #e2e8f0}
th:first-child{text-align:left}
td{padding:5px 7px;text-align:right;border-bottom:1px solid #f1f5f9}
td:first-child{text-align:left;max-width:180px;overflow:hidden;text-overflow:ellipsis}
tr:hover td{background:#f8fafc}
tr.l td{background:#fffbeb}
.tr td{font-weight:700;background:#f1f5f9;border-top:2px solid #94a3b8}
#ld{display:none;text-align:center;padding:40px;color:#666}
#ld .sp{display:inline-block;width:26px;height:26px;border:3px solid #e2e8f0;border-top-color:#2563eb;border-radius:50%;animation:spin .8s linear infinite;margin-bottom:10px}
@keyframes spin{to{transform:rotate(360deg)}}
#ra{display:none}
.er{color:#ef4444;background:#fef2f2;padding:12px;border-radius:6px;margin-bottom:12px}
.tp{font-size:12px;color:#94a3b8;margin-top:12px}
</style>
</head>
<body>
<div class="container">
<h1>📊 链接盈亏分析工具</h1>
<div class="ua">
  <div class="ci" id="ci"><span class="b">⏳</span><span>正在获取成本数据...</span></div>

  <div class="ur">
    <div class="ui">
      <label>1 订单数据</label>
      <div class="h">拼多多导出,含:商品id、商家编码-规格维度、商家实收</div>
      <input type="file" accept=".csv,.xls,.xlsx" id="of" onchange="upload('orders')">
      <div class="s" id="os">未上传</div>
    </div>
    <div class="ui">
      <label>2 推广报表</label>
      <div class="h">拼多多导出,含:商品ID、总花费(元)</div>
      <input type="file" accept=".csv,.xls,.xlsx" id="pf" onchange="upload('promo')">
      <div class="s" id="ps">未上传</div>
    </div>
  </div>
  <div class="br">
    <button class="btn bp" id="bg" onclick="analyze()" disabled>🚀 开始分析</button>
    <button class="btn bo" onclick="resetAll()">重置</button>
  </div>
</div>
<div id="ld"><div class="sp"></div><div>分析中...</div></div>
<div id="ra">
  <div class="sum" id="sc"></div>
  <div class="cs"><div style="font-size:14px;margin-bottom:8px;font-weight:600">费用构成</div><div class="cg" id="cg"></div></div>
  <div id="wn"></div>
  <div class="tw"><table><thead id="th"></thead><tbody id="tb"></tbody></table></div>
  <div class="br" style="margin-top:12px"><button class="btn bo" onclick="exportCSV()">📥 导出 CSV</button><span class="tp">数据仅在浏览器本地处理</span></div>
</div>
<div id="ar"></div>
<div id="er" class="er" style="display:none"></div>
</div>

<script>
var PUTIAN=[], YIWU=[], PM={}, YM={}, orders=null, promo=null, results=null, QTY_COL='';

// 页面加载后获取成本
fetch('/api/cost').then(function(r){return r.json();}).then(function(d){
  PUTIAN=d.p||[]; YIWU=d.y||[];
  PUTIAN.forEach(function(i){PM[i[0]]=i[1];});
  YIWU.forEach(function(i){YM[i[0]]=i[1];});
  document.getElementById('ci').innerHTML='<span class="b">✅</span><span>成本已加载:莆田 '+PUTIAN.length+' 项 · 义乌 '+YIWU.length+' 项</span>';
  checkReady();
}).catch(function(e){
  document.getElementById('ci').innerHTML='<span class="b" style="background:#ef4444">❌</span><span>成本获取失败:'+e.message+'</span>';
});

function st(id,t,c){var e=document.getElementById(id);e.textContent=t;e.style.color=c||'#999';}

function upload(type){
  var fid=(type=='orders'?'of':'pf'), sid=(type=='orders'?'os':'ps');
  var f=document.getElementById(fid).files[0];
  if(!f){st(sid,'未上传');return;}
  st(sid,'上传解析中...','#f59e0b');
  var fd=new FormData();
  fd.append('file',f);
  fd.append('type',type);
  fetch('/api/upload',{method:'POST',body:fd}).then(function(r){return r.json();}).then(function(d){
    if(d.ok){
      if(type=='orders'){
        orders=d.data;
        QTY_COL=d.qty_col||'';
        var qtyInfo=QTY_COL?' 数量列=['+QTY_COL+']':' ⚠️ 未检测到数量列';
        st(sid,'✅ '+d.count+' 条'+qtyInfo,'#22c55e');
      }else{
        promo=d.data;
        st(sid,'✅ '+d.count+' 条','#22c55e');
      }
      checkReady();
    }else{
      st(sid,'❌ '+d.msg,'#ef4444');
    }
  }).catch(function(e){st(sid,'❌ '+e.message,'#ef4444');});
}

function checkReady(){
  document.getElementById('bg').disabled=!(orders&&promo);
}

function resetAll(){
  orders=promo=results=null;
  document.getElementById('of').value=''; document.getElementById('pf').value='';
  st('os','未上传'); st('ps','未上传');
  document.getElementById('bg').disabled=true;
  document.getElementById('ra').style.display='none';
  document.getElementById('er').style.display='none';
  document.getElementById('ld').style.display='none';
}

function sm(s){
  if(!s) return ['',1];
  // 多盒: * X x 作为数量分隔符
  var m=s.match(/[\*Xx](\d+)$/);
  if(m) return [s.slice(0,m.index).trim(),parseInt(m[1])];
  return [s,1];
}
function mc(s){
  var p=sm(s),b=p[0],m=p[1];
  // 组合装:+ 连接,拆分逐项求和,保留Xx乘数
  if(b.indexOf('+')>=0){
    var parts=b.split('+'), total=0, ok=true;
    for(var i=0;i<parts.length;i++){
      var part=parts[i].trim();
      var xm=part.match(/[\*Xx](\d+)$/);
      var code=xm?part.slice(0,xm.index).trim():part;
      var mult=xm?parseInt(xm[1]):1;
      // 优先级:莆田PM(原始) → 义乌YM(去前导0) → 义乌YM(原始)
      if(PM[code]!==undefined) total+=PM[code]*mult;
      else {
        var b2=code.replace(/^0+/,'');
        if(YM[b2]!==undefined) total+=YM[b2]*mult;
        else if(YM[code]!==undefined) total+=YM[code]*mult;
        else { ok=false; break; }
      }
    }
    if(ok) return {c:total, m:1, s:'义乌'};
    return null;
  }
  // 单商品:优先级 莆田PM(原始) → 义乌YM(去前导0) → 义乌YM(原始)
  if(PM[b]!==undefined) return {c:PM[b],m:m,s:'莆田'};
  var b2=b.replace(/^0+/,'');
  if(YM[b2]!==undefined) return {c:YM[b2],m:m,s:'义乌'};
  if(YM[b]!==undefined) return {c:YM[b],m:m,s:'义乌'};
  return null;
}function analyze(){
  var ld=document.getElementById('ld'), ra=document.getElementById('ra'), er=document.getElementById('er');
  ra.style.display='none'; er.style.display='none'; ld.style.display='flex';
  setTimeout(function(){
    try{
      var match_count=0, nomatch_count=0, filtered_count=0, nocode_count=0, cost_map={}, unmatched=[];
      // 从推广报表构建商品名称映射
      var name_map={};
      for(var ni=0;ni<promo.length;ni++){
        var np_id=String(promo[ni]['商品ID']||'').trim();
        var np_name=promo[ni]['商品名称']||'';
        if(np_id&&np_name) name_map[np_id]=np_name;
      }
      for(var i=0;i<orders.length;i++){
        var o=orders[i];
        var code=o['商家编码-规格维度']||o['规格维度']||'';
        if(!code||code==='0'){ nocode_count++; nomatch_count++; unmatched.push({pid:o['商品id']||o['商品ID']||'',code:'(无编码)',name:name_map[o['商品id']||o['商品ID']||'']||o['商品名称']||'',income:parseFloat(o['商家实收金额(元)']||o['商家实收']||o['实收金额']||0),reason:'商家编码为空'}); continue; }
        var status=o['订单状态']||o['状态']||'';
        var skip_statuses=['未发货','退款成功','待付款','已取消'];
        var is_skip=false;
        for(var si=0;si<skip_statuses.length;si++){ if(status.indexOf(skip_statuses[si])>=0){ is_skip=true; break; } }
        if(is_skip){ filtered_count++; nomatch_count++; unmatched.push({pid:o['商品id']||o['商品ID']||'',code:code,name:name_map[o['商品id']||o['商品ID']||'']||o['商品名称']||o['商品标题']||o['名称']||'',income:parseFloat(o['商家实收金额(元)']||o['商家实收']||o['实收金额']||0),reason:'订单状态('+status+')'}); continue; }
        var mc_result=mc(code);
        if(mc_result){
          match_count++;
          var rowQty=parseInt(o['商品数量']||o['购买数量']||o['数量']||1);var cost=mc_result.c*mc_result.m*rowQty;
          var income=parseFloat(o['商家实收金额(元)']||o['商家实收']||o['实收金额']||0);
          var pack_cost=cost*0.05;
          var tech_fee=income*0.006;
          var tax=income*0.015;
          var refund=income*0.05;
          var pid=o['商品id']||o['商品ID']||'';
          var item={pid:pid,code:code,name:name_map[pid]||o['商品名称']||o['商品标题']||o['名称']||'',spec:o['规格']||o['规格名称']||'',income:income,cost:cost,pack_cost:pack_cost,tech_fee:tech_fee,tax:tax,refund:refund,warehouse:mc_result.s};
          if(pid&&!cost_map[pid]) cost_map[pid]={count:0,income:0,cost:0,pack_cost:0,tech_fee:0,tax:0,refund:0,orders:[],priceErrorCount:0};
          cost_map[pid].count++;
          cost_map[pid].income+=income;
          cost_map[pid].cost+=cost;
          cost_map[pid].pack_cost+=pack_cost;
          cost_map[pid].tech_fee+=tech_fee;
          cost_map[pid].tax+=tax;
          cost_map[pid].refund+=refund;
          cost_map[pid].orders.push(item);
          if(income < cost + 2) cost_map[pid].priceErrorCount++;
        }else{ nomatch_count++; unmatched.push({pid:o['商品id']||o['商品ID']||'',code:code,name:name_map[o['商品id']||o['商品ID']||'']||o['商品名称']||o['商品标题']||o['名称']||'',income:parseFloat(o['商家实收金额(元)']||o['商家实收']||o['实收金额']||0),reason:'成本未匹配'}); }
      }

      // 推广费:逐行累加,跳过合计行(商品ID为合计/总计)
      var promo_cost_map={};
      for(var pi=0;pi<promo.length;pi++){
        var pid2=String(promo[pi]['商品ID']||'').trim();
        var pid2_name=String(promo[pi]['商品名称']||'').trim();
        if(pid2.indexOf('计')>=0||pid2===''||pid2_name.indexOf('计')>=0) continue; // 跳过合计行
        var spend_str=promo[pi]['总花费(元)'];
        if(spend_str===undefined||spend_str==='') spend_str=promo[pi]['成交花费(元)'];
        var spend=parseFloat(spend_str||0);
        if(spend>0) promo_cost_map[pid2]=(promo_cost_map[pid2]||0)+spend;
      }
      // 统计无订单链接的推广费
      var orphan_promo=[];
      for(var pid2 in promo_cost_map) if(!cost_map[pid2]) orphan_promo.push({pid:pid2,promo:promo_cost_map[pid2]});
      var orphan_total=orphan_promo.reduce(function(s,i){return s+i.promo;},0);
      var total_income=0,total_cost=0,total_pack=0,total_tech=0,total_tax=0,total_refund=0,total_delivery=0;
      var total_by_order_promo=0;
      for(var key in cost_map){
        var p=cost_map[key];
        p.promo=promo_cost_map[key]||0;
        p.delivery=p.count*2.5;
        var r1v=function(v){return Math.round(v*10)/10;};
        p.net_profit=r1v(p.income-p.cost-p.pack_cost-p.promo-p.tech_fee-p.delivery-p.tax-p.refund);
        total_income+=p.income; total_cost+=p.cost; total_pack+=p.pack_cost;
        total_tech+=p.tech_fee; total_tax+=p.tax; total_refund+=p.refund;
        total_by_order_promo+=p.promo; total_delivery+=p.delivery;
      }
      var total_promo=total_by_order_promo+orphan_total;

      var net_total=r1(total_income-total_cost-total_pack-total_promo-total_tech-total_delivery-total_tax-total_refund);
      var rate_total=total_income>0?r1(net_total/total_income*100):0;

      // 汇总
      st('sc','成交额 '+r1(total_income)+' 元 · 成本 '+r1(total_cost)+' 元 · 推广费 '+r1(total_promo)+' 元 · 净利润 '+r1(net_total)+' 元 ('+rate_total+'%)','#333');

      var cg='<table>';
      cg+='<tr><td>成交额</td><td>'+r1(total_income)+'</td></tr>';
      cg+='<tr><td>货款成本</td><td>-'+r1(total_cost)+'</td></tr>';
      cg+='<tr><td>包装损耗(5%)</td><td>-'+r1(total_pack)+'</td></tr>';
      cg+='<tr><td>推广费</td><td>-'+r1(total_promo)+'</td></tr>';
      cg+='<tr><td>技术服务费(0.6%)</td><td>-'+r1(total_tech)+'</td></tr>';
      cg+='<tr><td>快递费</td><td>-'+r1(total_delivery)+' ('+Math.round(total_delivery/2.5)+'单×2.5)</td></tr>';
      cg+='<tr><td>税费(1.5%)</td><td>-'+r1(total_tax)+'</td></tr>';
      cg+='<tr><td>退款损耗(5%)</td><td>-'+r1(total_refund)+'</td></tr>';
      cg+='<tr style="font-weight:700;color:'+(net_total>=0?'#22c55e':'#ef4444')+'"><td>净利润</td><td>'+r1(net_total)+'</td></tr>';
      cg+='<tr><td>利润率</td><td>'+rate_total+'%</td></tr>';
      document.getElementById('cg').innerHTML=cg;

      // 明细表
      var th='<tr><th>序号</th><th>商品ID</th><th>商品名称</th><th>规格</th><th>商家编码</th><th>仓库</th><th>成交额</th><th>货款成本</th><th>包装损耗</th><th>推广费</th><th>技术服务费</th><th>快递费</th><th>税费</th><th>退款损耗</th><th>净利润</th><th>利润率</th><th>订单数</th></tr>';
      var tb='', idx=0, wp_cnt=0, yw_cnt=0, loss_rows=[];
      for(var key in cost_map){
        var p=cost_map[key]; idx++;
        var is_yi=p.orders[0].warehouse==='义乌';
        if(is_yi) yw_cnt+=p.count; else wp_cnt+=p.count;
        var wr=is_yi?'义乌':'莆田';
        var oo=p.orders[0];
        var np=p.net_profit;
        var pr=p.income>0?r1(np/p.income*100):0;
        if(np<-20) loss_rows.push({pid:key,name:oo.name,net_profit:np,income:p.income,cost:p.cost,count:p.count});
        tb+='<tr'+(np<0?' class="l"':'')+'><td>'+idx+'</td><td>'+e(key)+'</td><td>'+e(oo.name)+'</td><td>'+e(oo.spec||'')+'</td><td>'+e(oo.code)+'</td><td>'+wr+'</td><td>'+r1(p.income)+'</td><td>'+r1(p.cost)+'</td><td>'+r1(p.pack_cost)+'</td><td>'+r1(p.promo)+'</td><td>'+r1(p.tech_fee)+'</td><td>'+r1(p.delivery)+'</td><td>'+r1(p.tax)+'</td><td>'+r1(p.refund)+'</td><td style="color:'+(np>=0?'#22c55e':'#ef4444')+'">'+r1(np)+'</td><td>'+pr+'%</td><td>'+p.count+'</td></tr>';
      }
      document.getElementById('th').innerHTML=th;
      document.getElementById('tb').innerHTML=tb;

      // 匹配汇总
      document.getElementById('wn').innerHTML='<div style="font-size:14px;margin-bottom:8px;font-weight:600">匹配结果</div>'
        +'<div style="margin-bottom:8px">匹配 <span style="color:#22c55e">'+match_count+'</span> 条 · 未匹配 <span style="color:#ef4444">'+nomatch_count+'</span> 条'
        +'<span style="font-size:11px;color:#94a3b8;margin-left:8px">(已过滤 '+filtered_count+' 条 · 无编码 '+nocode_count+' 条 · 成本未匹配 '+(nomatch_count-filtered_count-nocode_count)+' 条)</span></div>'
        +'<div>莆田 <span style="color:#2563eb">'+wp_cnt+'</span> 条 · 义乌 <span style="color:#2563eb">'+yw_cnt+'</span> 条</div>';

      // 价格错误提醒
      var priceErrorRows=[];
      for(var peKey in cost_map){
        var peP=cost_map[peKey];
        if(peP.priceErrorCount>0 && peP.orders[0]) priceErrorRows.push({pid:peKey,name:peP.orders[0].name,count:peP.priceErrorCount});
      }
      // ---------- 分析报告 ----------
      var ar='<div style="background:#fff;border-radius:10px;padding:14px;margin-top:16px;box-shadow:0 1px 4px rgba(0,0,0,0.08)">';
      ar+='<div style="font-size:14px;margin-bottom:10px;font-weight:600">📋 分表数据分析报告</div>';

      // 亏损>20
      if(loss_rows.length>0){
        loss_rows.sort(function(a,b){return a.net_profit-b.net_profit;});
        ar+='<div style="margin-bottom:8px;font-weight:600;color:#ef4444">🔴 亏损>20元的链接 ('+loss_rows.length+'个)</div>';
        ar+='<table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:12px">';
        ar+='<tr><th>#</th><th>商品ID</th><th>商品名称</th><th>成交额</th><th>成本</th><th>净利润</th><th>订单数</th></tr>';
        for(var li=0;li<loss_rows.length;li++){
          var l=loss_rows[li];
          ar+='<tr class="l"><td>'+(li+1)+'</td><td>'+e(l.pid)+'</td><td>'+e(l.name)+'</td><td>'+r1(l.income)+'</td><td>'+r1(l.cost)+'</td><td style="color:#ef4444">'+r1(l.net_profit)+'</td><td>'+l.count+'</td></tr>';
        }
        ar+='</table>';
      }else{
        ar+='<div style="margin-bottom:8px;color:#22c55e">✅ 没有亏损超过20元的链接</div>';
      }

      // 价格错误提醒
      if(priceErrorRows.length>0){
        ar+='<div style="margin-bottom:8px;font-weight:600;color:#dc2626">🔴 价格错误（成交价<成本+2元）('+priceErrorRows.length+'个)</div>';
        for(var pei=0;pei<priceErrorRows.length;pei++){
          var pe=priceErrorRows[pei];
          ar+='<div style="padding:6px 10px;margin-bottom:4px;background:#fef2f2;border-radius:6px;font-size:13px">[下架] 商品ID:'+e(pe.pid)+' '+e(pe.name)+' 出'+pe.count+'单 - 该产品价格错误，请注意检查</div>';
        }
      }

      // 无订单链接的推广费
      if(orphan_promo.length>0){
        orphan_promo.sort(function(a,b){return b.promo-a.promo;});
        ar+='<div style="margin-bottom:8px;font-weight:600;color:#f59e0b">🟡 有推广无订单的链接 ('+orphan_promo.length+'个, 合计'+r1(orphan_total)+'元)</div>';
        ar+='<table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:12px">';
        ar+='<tr><th>#</th><th>商品ID</th><th>推广花费</th></tr>';
        for(var oi=0;oi<orphan_promo.length;oi++){
          var op=orphan_promo[oi];
          ar+='<tr><td>'+(oi+1)+'</td><td>'+e(op.pid)+'</td><td style="color:#ef4444">'+r1(op.promo)+'</td></tr>';
        }
        ar+='</table>';
      }

      ar+='<div style="font-size:13px;margin-bottom:8px;font-weight:600">📖 字段说明</div>';
      ar+='<table style="width:100%;border-collapse:collapse;font-size:11px;margin-bottom:12px">';
      ar+='<tr><th>字段</th><th>说明</th></tr>';
      ar+='<tr><td>成交额</td><td>订单中商家实收金额(元)之和</td></tr>';
      ar+='<tr><td>货款成本</td><td>按商家编码匹配莆田/义乌成本表,包含Xx多盒乘数</td></tr>';
      ar+='<tr><td>包装损耗</td><td>货款成本 × 5%</td></tr>';
      ar+='<tr><td>推广费</td><td>推广报表各商品ID的总花费(元)之和（跳过合计行），含无订单链接</td></tr>';
      ar+='<tr><td>技术服务费</td><td>成交额 × 0.6%(拼多多平台收取)</td></tr>';
      ar+='<tr><td>快递费</td><td>订单数 × 2.5元/单</td></tr>';
      ar+='<tr><td>税费</td><td>成交额 × 1.5%</td></tr>';
      ar+='<tr><td>退款损耗</td><td>成交额 × 5%(预估)</td></tr>';
      ar+='<tr><td>净利润</td><td>成交额-货款成本-包装损耗-推广费-技术服务费-快递费-税费-退款损耗</td></tr>';
      ar+='</table>';

      // 未匹配数据明细
      if(unmatched.length>0){
        ar+='<div style="font-size:13px;margin-bottom:8px;font-weight:600;margin-top:12px">📄 未匹配订单明细 ('+unmatched.length+'条)</div>';
        ar+='<div style="max-height:300px;overflow-y:auto;margin-bottom:12px">';
        ar+='<table style="width:100%;border-collapse:collapse;font-size:11px">';
        ar+='<tr><th>商品ID</th><th>商家编码</th><th>商品名称</th><th>成交额</th><th>原因</th></tr>';
        for(var ui=0;ui<unmatched.length;ui++){
          var u=unmatched[ui];
          ar+='<tr><td>'+e(u.pid)+'</td><td>'+e(u.code)+'</td><td>'+e(u.name)+'</td><td>'+r1(u.income)+'</td><td style="color:#ef4444;font-size:10px">'+u.reason+'</td></tr>';
        }
        ar+='</table></div>';
      }

      // 分析建议
      ar+='<div style="font-size:13px;margin-bottom:8px;font-weight:600">💡 分析建议</div>';
      var sug=[];
      if(net_total<0) sug.push('整体净利润为负('+r1(net_total)+'元),需重点关注:检查成本端是否有降价空间、调整亏损链接的售价或暂停推广');
      else if(rate_total<5) sug.push('整体利润率仅'+rate_total+'%,偏低,建议逐一排查亏损链接的原因');
      else sug.push('整体利润率'+rate_total+'%,表现良好,持续关注亏损链接');

      if(loss_rows.length>0){
        var loss_amt=loss_rows.reduce(function(s,l){return s+Math.abs(l.net_profit);},0);
        sug.push('亏损>20元的链接共'+loss_rows.length+'个,累计亏损'+r1(loss_amt)+'元。建议:1检查这些链接的推广花费是否过高 2确认成本匹配是否正确 3考虑暂停或调价');
        var top3=loss_rows.slice(0,Math.min(3,loss_rows.length));
        sug.push('亏损TOP'+top3.length+':'+top3.map(function(l){return e(l.pid)+'('+r1(l.net_profit)+'元)';}).join('、'));
      }

      var promo_ratio=total_income>0?r1(total_promo/total_income*100):0;
      if(promo_ratio>15) sug.push('推广费占成交额'+promo_ratio+'%,偏高。检查各链接投产比,ROI<1的链接建议暂停或降低出价');
      else if(promo_ratio<3) sug.push('推广费仅占成交额'+promo_ratio+'%,偏低,可考虑适当增加推广获取更多流量');
      else sug.push('推广费占成交额'+promo_ratio+'%,在合理范围内');

      ar+='<ul style="font-size:12px;line-height:1.8;margin:0;padding-left:18px">';
      for(var si=0;si<sug.length;si++) ar+='<li>'+sug[si]+'</li>';
      ar+='</ul></div>';

      document.getElementById('ar').innerHTML=ar;

      ra.style.display='block';
      ld.style.display='none';
    }catch(ex){
      er.innerHTML='❌ 分析出错:'+ex.message;
      er.style.display='block';
      ld.style.display='none';
    }
  },50);
}function r1(v){return Math.round(v*10)/10;}
function e(s){return String(s).replace(/</g,'&lt;').replace(/>/g,'&gt;');}

function exportCSV(){
  var cost_map={}, unmatched_ex=[];
  // 从推广报表构建商品名称映射
  var name_map={};
  for(var ni=0;ni<promo.length;ni++){
    var np_id=String(promo[ni]['商品ID']||'').trim();
    var np_name=promo[ni]['商品名称']||'';
    if(np_id&&np_name) name_map[np_id]=np_name;
  }
  for(var i=0;i<orders.length;i++){
    var o=orders[i];
    var code=o['商家编码-规格维度']||o['规格维度']||'';
    var pid_ex=o['商品id']||o['商品ID']||'';
    var income_ex=parseFloat(o['商家实收金额(元)']||o['商家实收']||o['实收金额']||0);
    var name_ex=name_map[pid_ex]||o['商品名称']||o['商品标题']||o['名称']||'';
    if(!code||code==='0'){ unmatched_ex.push({pid:pid_ex,code:'(无编码)',name:name_ex,income:income_ex,reason:'无编码'}); continue; }
    var status=o['订单状态']||o['状态']||'';
    var skip_statuses=['未发货','退款成功','待付款','已取消'];
    var is_skip=false;
    for(var si=0;si<skip_statuses.length;si++){ if(status.indexOf(skip_statuses[si])>=0){ is_skip=true; break; } }
    if(is_skip){ unmatched_ex.push({pid:pid_ex,code:code,name:name_ex,income:income_ex,reason:'订单状态('+status+')'}); continue; }
    var mc_result=mc(code);
    if(!mc_result){ unmatched_ex.push({pid:pid_ex,code:code,name:name_ex,income:income_ex,reason:'成本未匹配'}); continue; }
    var income=income_ex;
    var rowQty=parseInt(o['商品数量']||o['购买数量']||o['数量']||1);var cost=mc_result.c*mc_result.m*rowQty;
    var pack_cost=cost*0.05;
    var tech_fee=income*0.006;
    var tax=income*0.015;
    var refund=income*0.05;
    var pid=o['商品id']||o['商品ID']||'';
    if(!pid||!cost_map[pid]) cost_map[pid]={orders:[],income:0,cost:0,pack_cost:0,tech_fee:0,tax:0,refund:0,name:name_map[pid]||o['商品名称']||o['商品标题']||o['名称']||'',code:code,warehouse:mc_result.s,priceErrorCount:0};
    cost_map[pid].income+=income;
    cost_map[pid].cost+=cost;
    cost_map[pid].pack_cost+=pack_cost;
    cost_map[pid].tech_fee+=tech_fee;
    cost_map[pid].tax+=tax;
    cost_map[pid].refund+=refund;
    cost_map[pid].orders.push(o);
    if(income < cost + 2) cost_map[pid].priceErrorCount++;
  }
  var promo_cost_map_export={};
  for(var pi=0;pi<promo.length;pi++){
    var pid2=String(promo[pi]['商品ID']||'').trim();
    var pid2_name=String(promo[pi]['商品名称']||'').trim();
    if(pid2.indexOf('计')>=0||pid2===''||pid2_name.indexOf('计')>=0) continue;
    var spend_str=promo[pi]['总花费(元)'];
    if(spend_str===undefined||spend_str==='') spend_str=promo[pi]['成交花费(元)'];
    var spend=parseFloat(spend_str||0);
    if(spend>0) promo_cost_map_export[pid2]=(promo_cost_map_export[pid2]||0)+spend;
  }
  var orphan_promo_export=[];
  for(var pid2 in promo_cost_map_export) if(!cost_map[pid2]) orphan_promo_export.push({pid:pid2,promo:promo_cost_map_export[pid2]});
  var orphan_total_export=orphan_promo_export.reduce(function(s,i){return s+i.promo;},0);
  var total_promo_export=0;
  for(var pid2 in promo_cost_map_export) total_promo_export+=promo_cost_map_export[pid2];
  function es(v){return String(v).replace(/,/g,'&#44;').replace(/"/g,'""');}
  var lines=['商品ID,商品名称,商家编码,仓库,成交额,货款成本,包装损耗,推广费,技术服务费,快递费,税费,退款损耗,净利润,利润率,订单数'];
  var ti=0,tc=0,tp2=0,tpr=0,tt=0,td=0,ttx=0,trf=0;
  for(var key in cost_map){
    var p=cost_map[key];
    p.promo=promo_cost_map_export[key]||0;
    p.delivery=p.orders.length*2.5;
    p.net_profit=r1(p.income-p.cost-p.pack_cost-p.promo-p.tech_fee-p.delivery-p.tax-p.refund);
    var rate=p.income>0?r1(p.net_profit/p.income*100):0;
    ti+=p.income;tc+=p.cost;tp2+=p.pack_cost;tpr+=p.promo;tt+=p.tech_fee;td+=p.delivery;ttx+=p.tax;trf+=p.refund;
    lines.push([es(key),es(p.name),es(p.code),p.warehouse,r1(p.income),r1(p.cost),r1(p.pack_cost),r1(p.promo),r1(p.tech_fee),r1(p.delivery),r1(p.tax),r1(p.refund),r1(p.net_profit),rate+'%',p.orders.length].join(','));
  }
  var loss_rows_export=0;
  for(var lkey in cost_map){
    var lp=cost_map[lkey];
    var lpp=promo_cost_map_export[lkey]||0;
    var ld2=lp.orders.length*2.5;
    var lnp=r1(lp.income-lp.cost-lp.pack_cost-lpp-lp.tech_fee-ld2-lp.tax-lp.refund);
    if(lnp<-20) loss_rows_export++;
  }
  var net_nop=r1(ti-tc-tp2-tpr-tt-td-ttx-trf);
  var net_w_orphan=r1(net_nop-orphan_total_export);
  var r2=ti>0?r1(net_w_orphan/ti*100)+'%':'';
  lines.push(['合计','','','',r1(ti),r1(tc),r1(tp2),r1(tpr),r1(tt),r1(td),r1(ttx),r1(trf),r1(net_w_orphan),r2,''].join(','));
  for(var oi2=0;oi2<orphan_promo_export.length;oi2++){
    var op2=orphan_promo_export[oi2];
    lines.push([es(op2.pid),'无订单','','',0,0,0,op2.promo,0,0,0,0,-op2.promo,'-',''].join(','));
  }
  lines.push(['','','','','','','','','','','','','','','']);
  lines.push(['=== 分析报告摘要 ===','','','','','','','','','','','','','','']);
  lines.push(['亏损>20链接: '+loss_rows_export+'个','','','','','','','','','','','','','','']);
  lines.push(['有推广无订单: '+orphan_promo_export.length+'个, 合计'+r1(orphan_total_export)+'元','','','','','','','','','','','','','','']);
  lines.push(['推广费总计(含无订单): '+r1(total_promo_export)+'元','','','','','','','','','','','','','','']);
  lines.push(['总净利润: '+r1(net_w_orphan)+'元','','','','','','','','','','','','','','']);
  // 价格错误提醒
  var priceErrorExport=[];
  for(var peKey2 in cost_map){
    var peP2=cost_map[peKey2];
    if(peP2.priceErrorCount>0 && peP2.orders[0]) priceErrorExport.push({pid:peKey2,name:peP2.orders[0].name,count:peP2.priceErrorCount});
  }
  for(var pei2=0;pei2<priceErrorExport.length;pei2++){
    var pe2=priceErrorExport[pei2];
    lines.push(['[下架] 商品ID:'+es(pe2.pid)+' '+es(pe2.name)+' 出'+pe2.count+'单 - 该产品价格错误，请注意检查','','','','','','','','','','','','','',''].join(','));
  }
  // 未匹配数据
  lines.push(['','','','','','','','','','','','','','','']);
  lines.push(['=== 未匹配订单明细 ('+unmatched_ex.length+'条) ===','','','','','','','','','','','','','','']);
  lines.push(['商品ID','商家编码','商品名称','成交额','原因','','','','','','','','','','']);
  for(var ui2=0;ui2<unmatched_ex.length;ui2++){
    var u2=unmatched_ex[ui2];
    lines.push([es(u2.pid),es(u2.code),es(u2.name),r1(u2.income),u2.reason,'','','','','','','','','',''].join(','));
  }
  var blob=new Blob(['\ufeff'+lines.join('\n')],{type:'text/csv;charset=utf-8'});
  var a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='\u94FE\u63A5\u76C8\u4E8F\u5206\u6790_'+new Date().toISOString().slice(0,10)+'.csv';a.click();
  URL.revokeObjectURL(a.href);
}</script>
</body>
</html>"""

# ========== 服务端 ==========
cost_data = None

class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/api/cost":
            self.send_json(200, cost_data)
        else:
            self.send_response(200)
            self.send_header("Content-Type", "text/html;charset=utf-8")
            self.end_headers()
            self.wfile.write(HTML.encode("utf-8"))

    def parse_multipart(self, body, content_type):
        """手动解析 multipart/form-data,不用 cgi(Python 3.14 已移除)"""
        boundary = ""
        for part in content_type.split(";"):
            p = part.strip()
            if p.startswith("boundary="):
                boundary = p[9:].strip('"')
        if not boundary:
            return None, None, ""

        boundary_bytes = ("--" + boundary).encode()
        parts = body.split(boundary_bytes)
        file_data = None
        file_type = None
        filename = ""

        for part in parts:
            if b"\r\n\r\n" in part:
                hdrs, data = part.split(b"\r\n\r\n", 1)
                hdr_text = hdrs.decode("utf-8", errors="replace")
                if 'name="file"' in hdr_text:
                    file_data = data.rstrip(b"\r\n--")
                    m = re.search(r'filename="([^"]*)"', hdr_text)
                    if m:
                        filename = m.group(1)
                elif 'name="type"' in hdr_text:
                    file_type = data.strip().decode("utf-8", errors="replace")
        return file_data, file_type, filename

    def do_POST(self):
        if self.path == "/api/upload":
            try:
                length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(length)
                content_type = self.headers.get("Content-Type", "")
                file_data, file_type, filename = self.parse_multipart(body, content_type)

                # DEBUG: 打印前200字节看看实际数据
                print("=== DEBUG UPLOAD ===")
                print("file_type:", file_type)
                print("filename:", filename)
                if file_data:
                    print("first 200 bytes:", file_data[:200])
                print("====================")

                if file_data is None:
                    self.send_json(400, {"ok": False, "msg": "文件数据解析失败,检查 multipart"})
                    return

                if file_type == "orders":
                    try:
                        rows = parse_any_file(file_data)
                        if rows:
                            keys=list(rows[0].keys())
                            print("=== ORDERS COLS ===")
                            for k in keys: print(f"  [{k}]")
                            print("===================")
                            print("订单状态样本:", repr(rows[0].get("订单状态", "NOT FOUND")))
                            print("状态样本:", repr(rows[0].get("状态", "NOT FOUND")))
                            print("商品名称样本:", repr(rows[0].get("商品名称", "NOT_FOUND")))
                            print("商品标题样本:", repr(rows[0].get("商品标题", "NOT_FOUND")))
                            print("名称样本:", repr(rows[0].get("名称", "NOT_FOUND")))
                        # 检测数量列
                        qty_col = ''
                        qty_keywords = ['商品数量','数量','商品件数','购买数量','件数','购买件数']
                        for hname in keys:
                            hclean = hname.replace(r'\(.*?\)','').replace(r'（.*?）','').strip()
                            for kw in qty_keywords:
                                if kw in hclean:
                                    qty_col = hname
                                    break
                            if qty_col:
                                break
                        print(f"检测到数量列: [{qty_col}]" if qty_col else "未检测到数量列")
                        self.send_json(200, {"ok": True, "count": len(rows), "data": rows, "keys": list(rows[0].keys()) if rows else [], "qty_col": qty_col})
                    except Exception as e:
                        traceback.print_exc()
                        self.send_json(400, {"ok": False, "msg": f"订单文件解析失败: {e}"})
                elif file_type == "promo":
                    try:
                        rows = parse_any_file(file_data)
                        self.send_json(200, {"ok": True, "count": len(rows), "data": rows})
                    except Exception as e:
                        traceback.print_exc()
                        self.send_json(400, {"ok": False, "msg": f"推广报表解析失败: {e}"})
                else:
                    self.send_json(400, {"ok": False, "msg": "未知文件类型"})
            except Exception as e:
                traceback.print_exc()
                self.send_json(500, {"ok": False, "msg": str(e)})
        else:
            self.send_json(404, {"ok": False, "msg": "not found"})

    def send_json(self, code, data):
        self.send_response(code)
        self.send_header("Content-Type", "application/json;charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def log_message(self, format, *args):
        pass

if __name__ == "__main__":
    import sys
    print = lambda *a: __builtins__.print(*a, flush=True)
    print("=" * 45)
    print("  链接盈亏分析工具 v3")
    print("  服务端解析文件 · 无需浏览器CDN")
    print("=" * 45)
    print()
    print("📦 获取成本数据...")
    cost_data = fetch_cost()
    print(f"   莆田 {len(cost_data['p'])} 项 · 义乌 {len(cost_data['y'])} 项")
    print()
    server = HTTPServer(("127.0.0.1", PORT), Handler)
    print(f"🌐 打开浏览器: http://127.0.0.1:{PORT}")
    webbrowser.open(f"http://127.0.0.1:{PORT}")
    print("   关闭本窗口即可停止服务")
    print()
    server.serve_forever()
